''' inventory additions (returned-broken products/materials) '''

from .models import (BrokenProducts, BrokenMaterials, ReturnedProducts, ReturnedMaterials)
from django.contrib.auth.decorators import login_required
from django.views.decorators.csrf import csrf_exempt
from django.http import JsonResponse, HttpResponse
from django.shortcuts import render
from openpyxl import Workbook


# ADD BROKEN-RETURNED PRODUCTS/MATERIALS PAGE VIEWS
''' The following function return view of broken products html view '''
@login_required
@csrf_exempt
def add_broken_products_view(request):
    return render(request, "inventory/additions/products/broken_add.html")

''' The following function return view of broken materials html view '''
@login_required
@csrf_exempt
def add_broken_materials_view(request):
    return render(request, "inventory/additions/materials/broken_add.html")

''' The following function return view of returned products html view '''
@login_required
@csrf_exempt
def add_returned_products_view(request):
    return render(request, "inventory/additions/products/returned_add.html")

''' The following function return view of returned materials html view '''
@login_required
@csrf_exempt
def add_returned_materials_view(request):
    return render(request, "inventory/additions/materials/returned_add.html")

# LIST OF BROKEN-RETURNED PRODUCTS/MATERIALS
''' The following function return broken products archive '''
@login_required
@csrf_exempt
def broken_products_list(request):
    if request.user.category == 4:
        products = BrokenProducts.objects.all().order_by('-product_date')
    elif request.user.category == 3:
        products = BrokenProducts.objects.all().order_by('-product_date')
    elif request.user.category == 2:
        products = BrokenProducts.objects.filter(product_location = 'مغازه غدیر').order_by('-product_date')
    elif request.user.category == 1:
        products = BrokenProducts.objects.filter(product_location = 'پلاک سه').order_by('-product_date')
    elif request.user.category == 0:
        products = BrokenProducts.objects.filter(product_location = 'انبار اخلاقی').order_by('-product_date')
    return render(request, "inventory/additions/products/broken_list.html", {'products' : products})

''' The following function return broken materials archive '''
@login_required
@csrf_exempt
def broken_materials_list(request):
    if request.user.category == 4:
        materials = BrokenMaterials.objects.all().order_by('-material_date')
    elif request.user.category == 3:
        materials = BrokenMaterials.objects.all().order_by('-material_date')
    elif request.user.category == 2:
        materials = BrokenMaterials.objects.filter(material_location = 'مغازه غدیر').order_by('-material_date')
    elif request.user.category == 1:
        materials = BrokenMaterials.objects.filter(material_location = 'پلاک سه').order_by('-material_date')
    elif request.user.category == 0:
        materials = BrokenMaterials.objects.filter(material_location = 'انبار اخلاقی').order_by('-material_date')
    return render(request, "inventory/additions/materials/broken_list.html", {'materials' : materials})


''' The following function return returned products archive '''
@login_required
@csrf_exempt
def returned_products_list(request):
    if request.user.category == 4:
        materials = ReturnedProducts.objects.all().order_by('-product_date')
    elif request.user.category == 3:
        materials = ReturnedProducts.objects.all().order_by('-product_date')
    elif request.user.category == 2:
        materials = ReturnedProducts.objects.filter(material_location = 'مغازه غدیر').order_by('-product_date')
    elif request.user.category == 1:
        materials = ReturnedProducts.objects.filter(material_location = 'پلاک سه').order_by('-product_date')
    elif request.user.category == 0:
        materials = ReturnedProducts.objects.filter(material_location = 'انبار اخلاقی').order_by('-product_date')
    return render(request, "inventory/additions/products/returned_list.html", {'materials' : materials})

''' The following function return returned materials archive '''
@login_required
@csrf_exempt
def returned_materials_list(request):
    if request.user.category == 4:
        materials = ReturnedMaterials.objects.all().order_by('-material_date')
    elif request.user.category == 3:
        materials = ReturnedMaterials.objects.all().order_by('-material_date')
    elif request.user.category == 2:
        materials = ReturnedMaterials.objects.filter(material_location = 'مغازه غدیر').order_by('-material_date')
    elif request.user.category == 1:
        materials = ReturnedMaterials.objects.filter(material_location = 'پلاک سه').order_by('-material_date')
    elif request.user.category == 0:
        materials = ReturnedMaterials.objects.filter(material_location = 'انبار اخلاقی').order_by('-material_date')
    return render(request, "inventory/additions/materials/returned_list.html", {'materials' : materials})

# REST SUBMIT NEW BROKEN-RETURNED PRODUCTS/MATERIALS
''' The following function submit new broken products '''
@login_required
@csrf_exempt
def js_add_broken_products(request):
    if request.method == 'POST':
        product_name = request.POST.get('product_name')
        product_code = request.POST.get('product_code')
        product_color = request.POST.get('product_color')
        product_location = request.POST.get('product_location')
        product_hall = request.POST.get('product_hall')
        product_unit = request.POST.get('product_unit')
        if product_name:
            if product_code:
                if product_color and product_color != "انتخاب رنگ":
                    if product_location and product_location != "انتخاب محل انبار":
                        if product_hall and product_hall != "انتخاب سالن انبار":
                            if product_unit and product_unit != "انتخاب واحد شمارش":
                                if BrokenProducts.objects.filter(product_location = product_location, product_code = product_code, product_color = product_color).exists():
                                    return JsonResponse({'status': 'محصولی با این رنگ بندی در این انبار موجود است', 'success': False})
                                else:
                                    full_name = request.user.first_name + " " + request.user.last_name
                                    BrokenProducts.objects.create(
                                        product_author = full_name,
                                        product_name = product_name,
                                        product_code = product_code,
                                        product_color = product_color,
                                        product_quantity = 0,
                                        product_location = product_location,
                                        product_hall = product_hall,
                                        product_unit = product_unit,
                                    )
                                    return JsonResponse({'status': 'محصول با موفقیت افزوده شد', 'success':True})
                            else:
                                return JsonResponse({'status': 'واحد شمارش انتخاب نشده', 'success': False})
                        else:
                            return JsonResponse({'status': 'سالن انبار انتخاب نشده', 'success': False})
                    else:
                        return JsonResponse({'status': 'محل انبار انتخاب نشده', 'success': False})
                else:
                    return JsonResponse({'status': 'رنگ محصول را وارد کنید', 'success': False})
            else:
                return JsonResponse({'status': 'کد محصول را وارد کنید', 'success': False})
        else:
            return JsonResponse({'status': 'نام محصول را وارد کنید', 'success': False})
    else:
        return JsonResponse({'status':'درخواست نامعتبر', 'success': False})

''' The following function submit new broken materials '''
@login_required
@csrf_exempt
def js_add_broken_materials(request):
    if request.method == 'POST':
        material_name = request.POST.get('material_name')
        material_code = request.POST.get('material_code')
        material_color = request.POST.get('material_color')
        material_location = request.POST.get('material_location')
        material_hall = request.POST.get('material_hall')
        material_unit = request.POST.get('material_unit')
        if material_name:
            if material_code:
                if material_color and material_color != "انتخاب رنگ":
                    if material_location and material_location != "انتخاب محل انبار":
                        if material_hall and material_hall != "انتخاب سالن انبار":
                            if material_unit and material_unit != "انتخاب واحد شمارش":
                                if BrokenMaterials.objects.filter(material_location = material_location, material_code = material_code, material_color = material_color).exists():
                                    return JsonResponse({'status': 'ماده اولیه با این رنگ بندی در این انبار موجود است', 'success': False})
                                else:
                                    full_name = request.user.first_name + " " + request.user.last_name
                                    BrokenMaterials.objects.create(
                                        material_author = full_name,
                                        material_name = material_name,
                                        material_code = material_code,
                                        material_color = material_color,
                                        material_quantity = 0,
                                        material_location = material_location,
                                        material_hall = material_hall,
                                        material_unit = material_unit,
                                    )
                                    return JsonResponse({'status': 'ماده اولیه با موفقیت افزوده شد', 'success':True})
                            else:
                                return JsonResponse({'status': 'واحد شمارش انتخاب نشده', 'success': False})
                        else:
                            return JsonResponse({'status': 'سالن انبار انتخاب نشده', 'success': False})
                    else:
                        return JsonResponse({'status': 'محل انبار انتخاب نشده', 'success': False})
                else:
                    return JsonResponse({'status': 'رنگ ماده اولیه را وارد کنید', 'success': False})
            else:
                return JsonResponse({'status': 'کد ماده اولیه را وارد کنید', 'success': False})
        else:
            return JsonResponse({'status': 'نام ماده اولیه را وارد کنید', 'success': False})
    else:
        return JsonResponse({'status':'درخواست نامعتبر', 'success': False})

''' The following function submit new returned products '''
@login_required
@csrf_exempt
def js_add_returned_products(request):
    if request.method == 'POST':
        product_name = request.POST.get('product_name')
        product_code = request.POST.get('product_code')
        product_color = request.POST.get('product_color')
        product_location = request.POST.get('product_location')
        product_hall = request.POST.get('product_hall')
        product_unit = request.POST.get('product_unit')
        if product_name:
            if product_code:
                if product_color and product_color != "انتخاب رنگ":
                    if product_location and product_location != "انتخاب محل انبار":
                        if product_hall and product_hall != "انتخاب سالن انبار":
                            if product_unit and product_unit != "انتخاب واحد شمارش":
                                if ReturnedProducts.objects.filter(product_location = product_location, product_code = product_code, product_color = product_color).exists():
                                    return JsonResponse({'status': 'محصولی با این رنگ بندی در این انبار موجود است', 'success': False})
                                else:
                                    full_name = request.user.first_name + " " + request.user.last_name
                                    ReturnedProducts.objects.create(
                                        product_author = full_name,
                                        product_name = product_name,
                                        product_code = product_code,
                                        product_color = product_color,
                                        product_quantity = 0,
                                        product_location = product_location,
                                        product_hall = product_hall,
                                        product_unit = product_unit,
                                    )
                                    return JsonResponse({'status': 'محصول با موفقیت افزوده شد', 'success':True})
                            else:
                                return JsonResponse({'status': 'واحد شمارش انتخاب نشده', 'success': False})
                        else:
                            return JsonResponse({'status': 'سالن انبار انتخاب نشده', 'success': False})
                    else:
                        return JsonResponse({'status': 'محل انبار انتخاب نشده', 'success': False})
                else:
                    return JsonResponse({'status': 'رنگ محصول را وارد کنید', 'success': False})
            else:
                return JsonResponse({'status': 'کد محصول را وارد کنید', 'success': False})
        else:
            return JsonResponse({'status': 'نام محصول را وارد کنید', 'success': False})
    else:
        return JsonResponse({'status':'درخواست نامعتبر', 'success': False})
    
''' The following function submit new returned materials '''
@login_required
@csrf_exempt
def js_add_returned_materials(request):
    if request.method == 'POST':
        material_name = request.POST.get('material_name')
        material_code = request.POST.get('material_code')
        material_color = request.POST.get('material_color')
        material_location = request.POST.get('material_location')
        material_hall = request.POST.get('material_hall')
        material_unit = request.POST.get('material_unit')
        if material_name:
            if material_code:
                if material_color and material_color != "انتخاب رنگ":
                    if material_location and material_location != "انتخاب محل انبار":
                        if material_hall and material_hall != "انتخاب سالن انبار":
                            if material_unit and material_unit != "انتخاب واحد شمارش":
                                if ReturnedMaterials.objects.filter(material_location = material_location, material_code = material_code, material_color = material_color).exists():
                                    return JsonResponse({'status': 'ماده اولیه با این رنگ بندی در این انبار موجود است', 'success': False})
                                else:
                                    full_name = request.user.first_name + " " + request.user.last_name
                                    ReturnedMaterials.objects.create(
                                        material_author = full_name,
                                        material_name = material_name,
                                        material_code = material_code,
                                        material_color = material_color,
                                        material_quantity = 0,
                                        material_location = material_location,
                                        material_hall = material_hall,
                                        material_unit = material_unit,
                                    )
                                    return JsonResponse({'status': 'ماده اولیه با موفقیت افزوده شد', 'success':True})
                            else:
                                return JsonResponse({'status': 'واحد شمارش انتخاب نشده', 'success': False})
                        else:
                            return JsonResponse({'status': 'سالن انبار انتخاب نشده', 'success': False})
                    else:
                        return JsonResponse({'status': 'محل انبار انتخاب نشده', 'success': False})
                else:
                    return JsonResponse({'status': 'رنگ ماده اولیه را وارد کنید', 'success': False})
            else:
                return JsonResponse({'status': 'کد ماده اولیه را وارد کنید', 'success': False})
        else:
            return JsonResponse({'status': 'نام ماده اولیه را وارد کنید', 'success': False})
    else:
        return JsonResponse({'status':'درخواست نامعتبر', 'success': False})

# EXCEL EXPORT LIST OF BROKEN-RETURNED PRODUCTS/MATERIALS
''' The following function return excel export for broken products list '''
@login_required
@csrf_exempt
def broken_products_list_excel_export(request):
    if request.user.category == 4:
        list = BrokenProducts.objects.all().order_by("product_date")
    elif request.user.category == 3:
        list = BrokenProducts.objects.all().order_by("product_date")
    elif request.user.category == 2:
        list = BrokenProducts.objects.filter(product_location = 'مغازه غدیر').order_by('product_date')
    elif request.user.category == 1:
        list = BrokenProducts.objects.filter(product_location = 'پلاک سه').order_by('product_date')
    elif request.user.category == 0:
        list = BrokenProducts.objects.filter(product_location = 'انبار اخلاقی').order_by('product_date')
    else:
        list = BrokenProducts.objects.all().order_by("product_date")
    wb = Workbook()
    ws = wb.active
    ws.append(["ردیف", "اقدام کننده", "نام محصول", "کد محصول", "رنگ محصول", "موجودی محصول", "انبار محصول", "سالن انبار محصول", "واحد محصول", "تاریخ ایجاد محصول", "فعال/غیرفعال", "موجود/ناموجود"])
    for data in list:
        ws.append([data.row, data.product_author, data.product_name, data.product_code, data.product_color, data.product_quantity, data.product_location, data.product_hall, data.product_unit, data.jpub(), data.is_active, data.is_available])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Products_lists.xlsx"
    wb.save(response)
    return response

''' The following function return excel export for broken materials list '''
@login_required
@csrf_exempt
def broken_materials_list_excel_export(request):
    if request.user.category == 4:
        list = BrokenMaterials.objects.all().order_by("material_date")
    elif request.user.category == 3:
        list = BrokenMaterials.objects.all().order_by("material_date")
    elif request.user.category == 2:
        list = BrokenMaterials.objects.filter(material_location = 'مغازه غدیر').order_by('material_date')
    elif request.user.category == 1:
        list = BrokenMaterials.objects.filter(material_location = 'پلاک سه').order_by('material_date')
    elif request.user.category == 0:
        list = BrokenMaterials.objects.filter(material_location = 'انبار اخلاقی').order_by('material_date')
    else:
        list = BrokenMaterials.objects.all().order_by("material_date")
    wb = Workbook()
    ws = wb.active
    ws.append(["ردیف", "اقدام کننده", "نام محصول", "کد محصول", "رنگ محصول", "موجودی محصول", "انبار محصول", "سالن انبار محصول", "واحد محصول", "تاریخ ایجاد محصول", "فعال/غیرفعال", "موجود/ناموجود"])
    for data in list:
        ws.append([data.row, data.material_author, data.material_name, data.material_code, data.material_color, data.material_quantity, data.material_location, data.material_hall, data.material_unit, data.jpub(), data.is_active, data.is_available])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Broken_lists.xlsx"
    wb.save(response)
    return response

''' The following function return excel export for returned products list '''
@login_required
@csrf_exempt
def returned_products_list_excel_export(request):
    if request.user.category == 4:
        list = ReturnedProducts.objects.all().order_by("product_date")
    elif request.user.category == 3:
        list = ReturnedProducts.objects.all().order_by("product_date")
    elif request.user.category == 2:
        list = ReturnedProducts.objects.filter(product_location = 'مغازه غدیر').order_by('product_date')
    elif request.user.category == 1:
        list = ReturnedProducts.objects.filter(product_location = 'پلاک سه').order_by('product_date')
    elif request.user.category == 0:
        list = ReturnedProducts.objects.filter(product_location = 'انبار اخلاقی').order_by('product_date')
    else:
        list = ReturnedProducts.objects.all().order_by("product_date")
    wb = Workbook()
    ws = wb.active
    ws.append(["ردیف", "اقدام کننده", "نام محصول", "کد محصول", "رنگ محصول", "موجودی محصول", "انبار محصول", "سالن انبار محصول", "واحد محصول", "تاریخ ایجاد محصول", "فعال/غیرفعال", "موجود/ناموجود"])
    for data in list:
        ws.append([data.row, data.product_author, data.product_name, data.product_code, data.product_color, data.product_quantity, data.product_location, data.product_hall, data.product_unit, data.jpub(), data.is_active, data.is_available])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Products_lists.xlsx"
    wb.save(response)
    return response

''' The following function return excel export for returned materials list '''
@login_required
@csrf_exempt
def returned_materials_list_excel_export(request):
    if request.user.category == 4:
        list = ReturnedMaterials.objects.all().order_by("material_date")
    elif request.user.category == 3:
        list = ReturnedMaterials.objects.all().order_by("material_date")
    elif request.user.category == 2:
        list = ReturnedMaterials.objects.filter(material_location = 'مغازه غدیر').order_by('material_date')
    elif request.user.category == 1:
        list = ReturnedMaterials.objects.filter(material_location = 'پلاک سه').order_by('material_date')
    elif request.user.category == 0:
        list = ReturnedMaterials.objects.filter(material_location = 'انبار اخلاقی').order_by('material_date')
    else:
        list = ReturnedMaterials.objects.all().order_by("material_date")
    wb = Workbook()
    ws = wb.active
    ws.append(["ردیف", "اقدام کننده", "نام محصول", "کد محصول", "رنگ محصول", "موجودی محصول", "انبار محصول", "سالن انبار محصول", "واحد محصول", "تاریخ ایجاد محصول", "فعال/غیرفعال", "موجود/ناموجود"])
    for data in list:
        ws.append([data.row, data.material_author, data.material_name, data.material_code, data.material_color, data.material_quantity, data.material_location, data.material_hall, data.material_unit, data.jpub(), data.is_active, data.is_available])
    response = HttpResponse(content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f"attachment; filename=Products_lists.xlsx"
    wb.save(response)
    return response

# PDF EXPORT LIST OF BROKEN-RETURNED PRODUCTS/MATERIALS
''' The following function return pdf export for broken products list '''
@login_required
@csrf_exempt
def broken_products_list_pdf_export(request):
    if request.user.category == 4:
        list = BrokenProducts.objects.all().order_by("product_date")
    elif request.user.category == 3:
        list = BrokenProducts.objects.all().order_by("product_date")
    elif request.user.category == 2:
        list = BrokenProducts.objects.filter(product_location = 'مغازه غدیر').order_by('product_date')
    elif request.user.category == 1:
        list = BrokenProducts.objects.filter(product_location = 'پلاک سه').order_by('product_date')
    elif request.user.category == 0:
        list = BrokenProducts.objects.filter(product_location = 'انبار اخلاقی').order_by('product_date')
    else:
        list = BrokenProducts.objects.all().order_by("product_date")
    context = {'title' : 'Materials list', 'list' : list}
    return render(request, 'utils/print_Broken_Products_list.html', context)

''' The following function return pdf export for broken materials list '''
@login_required
@csrf_exempt
def broken_materials_list_pdf_export(request):
    if request.user.category == 4:
        list = BrokenMaterials.objects.all().order_by("material_date")
    elif request.user.category == 3:
        list = BrokenMaterials.objects.all().order_by("material_date")
    elif request.user.category == 2:
        list = BrokenMaterials.objects.filter(material_location = 'مغازه غدیر').order_by('material_date')
    elif request.user.category == 1:
        list = BrokenMaterials.objects.filter(material_location = 'پلاک سه').order_by('material_date')
    elif request.user.category == 0:
        list = BrokenMaterials.objects.filter(material_location = 'انبار اخلاقی').order_by('material_date')
    else:
        list = BrokenMaterials.objects.all().order_by("material_date")
    context = {'title' : 'Materials list', 'list' : list}
    return render(request, 'utils/print_Broken_Materials_list.html', context)

''' The following function return pdf export for returned products list '''
@login_required
@csrf_exempt
def returned_products_list_pdf_export(request):
    if request.user.category == 4:
        list = ReturnedProducts.objects.all().order_by("product_date")
    elif request.user.category == 3:
        list = ReturnedProducts.objects.all().order_by("product_date")
    elif request.user.category == 2:
        list = ReturnedProducts.objects.filter(product_location = 'مغازه غدیر').order_by('product_date')
    elif request.user.category == 1:
        list = ReturnedProducts.objects.filter(product_location = 'پلاک سه').order_by('product_date')
    elif request.user.category == 0:
        list = ReturnedProducts.objects.filter(product_location = 'انبار اخلاقی').order_by('product_date')
    else:
        list = ReturnedProducts.objects.all().order_by("product_date")
    context = {'title' : 'Materials list', 'list' : list}
    return render(request, 'utils/print_Returned_Products_list.html', context)

''' The following function return pdf export for returned materials list '''
@login_required
@csrf_exempt
def returned_materials_list_pdf_export(request):
    if request.user.category == 4:
        list = ReturnedMaterials.objects.all().order_by("material_date")
    elif request.user.category == 3:
        list = ReturnedMaterials.objects.all().order_by("material_date")
    elif request.user.category == 2:
        list = ReturnedMaterials.objects.filter(material_location = 'مغازه غدیر').order_by('material_date')
    elif request.user.category == 1:
        list = ReturnedMaterials.objects.filter(material_location = 'پلاک سه').order_by('material_date')
    elif request.user.category == 0:
        list = ReturnedMaterials.objects.filter(material_location = 'انبار اخلاقی').order_by('material_date')
    else:
        list = ReturnedMaterials.objects.all().order_by("material_date")
    context = {'title' : 'Materials list', 'list' : list}
    return render(request, 'utils/print_Returned_Materials_list.html', context)